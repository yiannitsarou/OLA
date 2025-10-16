"""
step8_fixed_final.py
====================
Ολοκληρωμένο module για Fill + Optimize (χωρίς Streamlit UI).

Παρέχει όλη τη λειτουργικότητα του app.py ως library/CLI:
- Phase 1: Fill template με δεδομένα μαθητών
- Phase 2: Optimization με asymmetric swaps
- Locked students support (ΖΩΗΡΟΣ, ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ, ΙΔΙΑΙΤΕΡΟΤΗΤΑ)
- ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ + SINGLE sheets
- Detailed statistics + swaps log

Απαιτήσεις: openpyxl>=3.1.0
"""
from __future__ import annotations

import sys
import argparse
from pathlib import Path
from dataclasses import dataclass, field
from typing import Dict, List, Tuple, Optional, Any

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.worksheet.worksheet import Worksheet


# ========== DATACLASSES (FIXED για Python 3.12) ==========

@dataclass
class StudentData:
    """Δεδομένα μαθητή από source (Phase 1)."""
    name: str = ""
    gender: str = ""
    teacher_child: str = "Ο"
    calm: str = "Ο"
    special_needs: str = "Ο"
    greek_knowledge: str = "Ν"
    friends: List[str] = field(default_factory=list)
    conflicts: int = 0
    choice: int = 1


@dataclass
class Student:
    """Student για optimizer (Phase 2)."""
    name: str = ""
    choice: int = 1
    gender: str = ""
    greek_knowledge: str = "Ν"
    friends: List[str] = field(default_factory=list)
    locked: bool = False


# ========== MAIN PROCESSOR CLASS ==========

class UnifiedProcessor:
    """Ενοποιημένος processor: Fill + Optimize."""
    
    def __init__(self):
        self.students_data: Dict[str, StudentData] = {}
        self.teams_students: Dict[str, List[str]] = {}
        self.students: Dict[str, Student] = {}
        self.teams: Dict[str, List[str]] = {}
        self.target_ep3 = 3
        self.target_gender = 4
        self.target_greek = 4
        self.warnings: List[str] = []
    
    # ==================== PHASE 1: FILL ====================
    
    def read_source_data(self, source_path: str) -> None:
        """Διάβασμα δεδομένων από Παράδειγμα1.xlsx."""
        wb = load_workbook(source_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            headers = self._parse_headers_fill(sheet)
            
            if 'ΟΝΟΜΑ' not in headers:
                continue
            
            for row_idx in range(2, sheet.max_row + 1):
                name = self._get_cell_value(sheet, row_idx, headers.get('ΟΝΟΜΑ'))
                
                if not name:
                    continue
                
                # Friends
                friends_str = self._get_cell_value(sheet, row_idx, headers.get('ΦΙΛΟΙ'))
                friends = [f.strip() for f in friends_str.split(',') if f.strip()] if friends_str else []
                
                # Choice (ΕΠΙΔΟΣΗ)
                choice_val = 1
                if 'ΕΠΙΔΟΣΗ' in headers:
                    epidosi_raw = sheet.cell(row_idx, headers['ΕΠΙΔΟΣΗ']).value
                    if epidosi_raw is not None:
                        try:
                            choice_val = int(epidosi_raw)
                        except:
                            choice_val = 1
                
                # Greek knowledge - robust parsing
                greek_raw = None
                found_greek = False
                for variant in ['ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ', 'ΚΑΛΗ ΓΝΩΣΗ ΕΛΛΗΝΙΚΩΝ', 
                               'ΚΑΛΗ_ΓΝΩΣΗ', 'ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ']:
                    if variant in headers:
                        greek_raw = self._get_cell_value(sheet, row_idx, headers[variant], None)
                        if greek_raw is not None and greek_raw != '':
                            found_greek = True
                            break
                
                # Process Greek knowledge
                if not found_greek or greek_raw is None or greek_raw == '':
                    self.warnings.append(f"⚠️ Μαθητής {name}: Δεν βρέθηκε ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ - παραλείπεται")
                    continue
                else:
                    greek_str = str(greek_raw).strip().upper()
                    if greek_str.startswith('Ν') or greek_str.startswith('N'):
                        greek_val = 'Ν'
                    elif greek_str.startswith('Ο') or greek_str.startswith('O'):
                        greek_val = 'Ο'
                    else:
                        self.warnings.append(f"⚠️ Unknown ΚΑΛΗ_ΓΝΩΣΗ '{greek_raw}' for {name}, defaulting to Ν")
                        greek_val = 'Ν'
                
                self.students_data[name] = StudentData(
                    name=name,
                    gender=self._get_cell_value(sheet, row_idx, headers.get('ΦΥΛΟ'), 'Κ'),
                    teacher_child=self._get_cell_value(sheet, row_idx, headers.get('ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ'), 'Ο'),
                    calm=self._get_cell_value(sheet, row_idx, headers.get('ΖΩΗΡΟΣ'), 'Ο'),
                    special_needs=self._get_cell_value(sheet, row_idx, headers.get('ΙΔΙΑΙΤΕΡΟΤΗΤΑ'), 'Ο'),
                    greek_knowledge=greek_val,
                    friends=friends,
                    conflicts=0,
                    choice=choice_val
                )
        
        wb.close()
        print(f"✅ Διαβάστηκαν {len(self.students_data)} μαθητές από source file")
    
    def fill_target_excel(self, template_path: str, output_path: str) -> str:
        """Συμπλήρωση STEP7_TEMPLATE."""
        wb = load_workbook(template_path)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            filled_count = self._fill_sheet(sheet, sheet_name)
            if filled_count > 0:
                print(f"📝 Sheet '{sheet_name}': {filled_count} μαθητές")
        
        self._create_categorization_sheet(wb)
        
        wb.save(output_path)
        wb.close()
        
        print(f"✅ Filled Excel αποθηκεύτηκε: {output_path}")
        return output_path
    
    def _fill_sheet(self, sheet: Worksheet, team_name: str) -> int:
        """Συμπλήρωση ενός sheet."""
        headers_map = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                header = str(cell.value).strip().upper()
                header_key = header.replace('_', '').replace(' ', '')
                headers_map[header_key] = col_idx
        
        if 'ΟΝΟΜΑ' not in headers_map:
            return 0
        
        required_headers = ['ΦΥΛΟ', 'ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ', 'ΕΠΙΔΟΣΗ', 'ΦΙΛΟΙ']
        next_col = max(headers_map.values()) + 1
        
        for req_header in required_headers:
            if req_header not in headers_map:
                cell = sheet.cell(1, next_col)
                original_header = req_header.replace('ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ', 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ')
                cell.value = original_header
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(bold=True)
                headers_map[req_header] = next_col
                next_col += 1
        
        filled_count = 0
        
        if team_name not in self.teams_students:
            self.teams_students[team_name] = []
        
        for row_idx in range(2, sheet.max_row + 1):
            name_cell = sheet.cell(row_idx, headers_map['ΟΝΟΜΑ'])
            name = name_cell.value
            
            if not name or str(name).strip() == '':
                continue
            
            name = str(name).strip()
            
            if name not in self.students_data:
                continue
            
            student_data = self.students_data[name]
            self.teams_students[team_name].append(name)
            
            # Fill ΦΥΛΟ
            if 'ΦΥΛΟ' in headers_map:
                col = headers_map['ΦΥΛΟ']
                sheet.cell(row_idx, col).value = student_data.gender
                sheet.cell(row_idx, col).alignment = Alignment(horizontal='center', vertical='center')
            
            # Fill ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ
            if 'ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ' in headers_map:
                col = headers_map['ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ']
                sheet.cell(row_idx, col).value = student_data.greek_knowledge
                sheet.cell(row_idx, col).alignment = Alignment(horizontal='center', vertical='center')
            
            # Fill ΦΙΛΟΙ
            if 'ΦΙΛΟΙ' in headers_map:
                col = headers_map['ΦΙΛΟΙ']
                sheet.cell(row_idx, col).value = ', '.join(student_data.friends) if student_data.friends else ''
                sheet.cell(row_idx, col).alignment = Alignment(horizontal='left', vertical='center')
            
            # Fill ΕΠΙΔΟΣΗ
            if 'ΕΠΙΔΟΣΗ' in headers_map:
                col = headers_map['ΕΠΙΔΟΣΗ']
                sheet.cell(row_idx, col).value = student_data.choice
                sheet.cell(row_idx, col).alignment = Alignment(horizontal='center', vertical='center')
            
            filled_count += 1
        
        return filled_count
    
    def _create_categorization_sheet(self, workbook: Workbook) -> None:
        """Δημιουργία sheet ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ."""
        if 'ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ' in workbook.sheetnames:
            del workbook['ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ']
        
        cat_sheet = workbook.create_sheet('ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ')
        
        headers = ['ΜΑΘΗΤΗΣ Α', 'ΜΑΘΗΤΗΣ Β', 'ΚΑΤΗΓΟΡΙΑ ΔΥΑΔΑΣ', 'ΕΠΙΔΟΣΗ', 'LOCKED', 'ΤΜΗΜΑ']
        for col_idx, header in enumerate(headers, start=1):
            cell = cat_sheet.cell(1, col_idx)
            cell.value = header
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        all_students = []
        for team_name in sorted(self.teams_students.keys()):
            for student_name in self.teams_students[team_name]:
                if student_name in self.students_data:
                    student = self.students_data[student_name]
                    all_students.append({
                        'name': student_name,
                        'data': student,
                        'team': team_name
                    })
        
        row_idx = 2
        processed = set()
        
        for i, student_a in enumerate(all_students):
            if student_a['name'] in processed:
                continue
            
            for j, student_b in enumerate(all_students[i+1:], start=i+1):
                if student_b['name'] in processed:
                    continue
                
                if (student_b['name'] in student_a['data'].friends or 
                    student_a['name'] in student_b['data'].friends):
                    
                    category = self._determine_category(
                        student_a['data'].gender,
                        student_a['data'].greek_knowledge,
                        student_b['data'].gender,
                        student_b['data'].greek_knowledge
                    )
                    
                    epidosi_text = f"{student_a['data'].choice}, {student_b['data'].choice}"
                    
                    cat_sheet.cell(row_idx, 1).value = student_a['name']
                    cat_sheet.cell(row_idx, 2).value = student_b['name']
                    cat_sheet.cell(row_idx, 3).value = category
                    cat_sheet.cell(row_idx, 4).value = epidosi_text
                    
                    is_locked = (self._is_student_locked(student_a['data']) or 
                                 self._is_student_locked(student_b['data']))
                    cat_sheet.cell(row_idx, 5).value = 'LOCKED' if is_locked else 'ΟΧΙ'
                    
                    if is_locked:
                        team_text = 'LOCKED'
                    else:
                        team_text = f"{student_a['team']},{student_b['team']}"
                    cat_sheet.cell(row_idx, 6).value = team_text
                    
                    for col in range(1, 7):
                        cat_sheet.cell(row_idx, col).alignment = Alignment(
                            horizontal='left' if col <= 2 else 'center',
                            vertical='center'
                        )
                    
                    processed.add(student_a['name'])
                    processed.add(student_b['name'])
                    row_idx += 1
                    break
        
        cat_sheet.column_dimensions['A'].width = 30
        cat_sheet.column_dimensions['B'].width = 30
        cat_sheet.column_dimensions['C'].width = 35
        cat_sheet.column_dimensions['D'].width = 12
        cat_sheet.column_dimensions['E'].width = 12
        cat_sheet.column_dimensions['F'].width = 20
        
        self._create_single_sheet(workbook, all_students, processed)
    
    def _is_student_locked(self, student: StudentData) -> bool:
        """Έλεγχος αν μαθητής είναι locked."""
        return (student.calm == 'Ν' or 
                student.teacher_child == 'Ν' or 
                student.special_needs == 'Ν')
    
    def _determine_category(self, gender_a: str, greek_a: str, gender_b: str, greek_b: str) -> str:
        """Καθορισμός κατηγορίας δυάδας."""
        if gender_a != gender_b:
            return "Ομάδες Μικτού Φύλου"
        
        gender_label = "Κορίτσια" if gender_a == "Κ" else "Αγόρια"
        
        if greek_a == greek_b:
            if greek_a == "Ν":
                return f"Καλή Γνώση ({gender_label})"
            else:
                return f"όχι Καλή Γνώση ({gender_label})"
        else:
            return f"Μικτής Γνώσης ({gender_label})"
    
    def _determine_single_category(self, gender: str, greek_knowledge: str) -> str:
        """Καθορισμός κατηγορίας για μεμονωμένο μαθητή."""
        gender_label = "Κορίτσια" if gender == "Κ" else "Αγόρια"
        
        if greek_knowledge == "Ν":
            return f"{gender_label} - Ν (Καλή γνώση)"
        else:
            return f"{gender_label} - Ο (όχι καλή γνώση)"
    
    def _create_single_sheet(self, workbook: Workbook, all_students: List[Dict], processed_names: set) -> None:
        """Δημιουργία sheet SINGLE."""
        if 'SINGLE' in workbook.sheetnames:
            del workbook['SINGLE']
        
        single_sheet = workbook.create_sheet('SINGLE')
        
        headers = ['ΟΝΟΜΑ', 'ΦΥΛΟ', 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ', 'ΕΠΙΔΟΣΗ', 'ΚΑΤΗΓΟΡΙΑ SINGLE', 'ΤΜΗΜΑ', 'LOCKED']
        for col_idx, header in enumerate(headers, start=1):
            cell = single_sheet.cell(1, col_idx)
            cell.value = header
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        single_students = []
        for student in all_students:
            if student['name'] not in processed_names:
                single_students.append(student)
        
        single_students.sort(key=lambda x: x['name'])
        
        for row_idx, student in enumerate(single_students, start=2):
            student_data = student['data']
            team_name = student['team']
            
            single_sheet.cell(row_idx, 1).value = student['name']
            single_sheet.cell(row_idx, 2).value = student_data.gender
            single_sheet.cell(row_idx, 3).value = student_data.greek_knowledge
            single_sheet.cell(row_idx, 4).value = student_data.choice
            
            category = self._determine_single_category(student_data.gender, student_data.greek_knowledge)
            single_sheet.cell(row_idx, 5).value = category
            
            is_locked = self._is_student_locked(student_data)
            single_sheet.cell(row_idx, 7).value = 'LOCKED' if is_locked else 'ΟΧΙ'
            
            if is_locked:
                single_sheet.cell(row_idx, 6).value = 'LOCKED'
            else:
                single_sheet.cell(row_idx, 6).value = team_name
            
            for col in range(1, 8):
                single_sheet.cell(row_idx, col).alignment = Alignment(
                    horizontal='left' if col == 1 else 'center',
                    vertical='center'
                )
        
        single_sheet.column_dimensions['A'].width = 30
        single_sheet.column_dimensions['B'].width = 12
        single_sheet.column_dimensions['C'].width = 25
        single_sheet.column_dimensions['D'].width = 12
        single_sheet.column_dimensions['E'].width = 35
        single_sheet.column_dimensions['F'].width = 20
        single_sheet.column_dimensions['G'].width = 12
    
    # ==================== PHASE 2: OPTIMIZE ====================
    
    def load_filled_data(self, filled_path: str) -> None:
        """Φόρτωση δεδομένων από filled Excel για optimization."""
        wb = load_workbook(filled_path, data_only=True)
        
        if 'ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ' in wb.sheetnames:
            self._load_from_kategoriopoihsh(wb['ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ'])
        
        if 'SINGLE' in wb.sheetnames:
            self._load_from_single(wb['SINGLE'])
        
        for sheet_name in wb.sheetnames:
            if sheet_name in ['ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ', 'SINGLE']:
                continue
            
            sheet = wb[sheet_name]
            headers = self._parse_headers(sheet)
            
            if 'ΟΝΟΜΑ' not in headers:
                continue
            
            self.teams[sheet_name] = []
            
            for row_idx in range(2, sheet.max_row + 1):
                name = self._get_cell_value(sheet, row_idx, headers.get('ΟΝΟΜΑ'))
                if name and name in self.students:
                    self.teams[sheet_name].append(name)
        
        wb.close()
        print(f"✅ Φορτώθηκαν {len(self.students)} students, {len(self.teams)} teams")
    
    def _load_from_kategoriopoihsh(self, sheet: Worksheet) -> None:
        """Διάβασμα δυάδων από ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ sheet."""
        headers = self._parse_headers(sheet)
        
        required = ['ΜΑΘΗΤΗΣΑ', 'ΜΑΘΗΤΗΣΒ', 'ΚΑΤΗΓΟΡΙΑΔΥΑΔΑΣ', 'ΕΠΙΔΟΣΗ']
        missing = [h for h in required if h not in headers]
        if missing:
            return
        
        for row_idx in range(2, sheet.max_row + 1):
            name_a = self._get_cell_value(sheet, row_idx, headers.get('ΜΑΘΗΤΗΣΑ'))
            name_b = self._get_cell_value(sheet, row_idx, headers.get('ΜΑΘΗΤΗΣΒ'))
            category = self._get_cell_value(sheet, row_idx, headers.get('ΚΑΤΗΓΟΡΙΑΔΥΑΔΑΣ'))
            epidosh_raw = self._get_cell_value(sheet, row_idx, headers.get('ΕΠΙΔΟΣΗ'))
            locked_val = self._get_cell_value(sheet, row_idx, headers.get('LOCKED'))
            
            if not name_a or not name_b or not category:
                continue
            
            # Parse επίδοση
            epidosh_a, epidosh_b = 1, 1
            if ',' in epidosh_raw:
                parts = epidosh_raw.split(',')
                try:
                    epidosh_a = int(parts[0].strip())
                    epidosh_b = int(parts[1].strip())
                except:
                    pass
            
            # Get real data from students_data
            sa = self.students_data.get(name_a)
            sb = self.students_data.get(name_b)
            
            gender_a = sa.gender if sa else 'Α'
            gender_b = sb.gender if sb else 'Α'
            
            greek_a = sa.greek_knowledge if sa else 'Ν'
            greek_b = sb.greek_knowledge if sb else 'Ν'
            
            is_locked = (self._is_student_locked(sa) if sa else False) or \
                       (self._is_student_locked(sb) if sb else False)
            
            if name_a not in self.students:
                self.students[name_a] = Student(
                    name=name_a,
                    choice=epidosh_a,
                    gender=gender_a,
                    greek_knowledge=greek_a,
                    friends=[name_b],
                    locked=is_locked
                )
            
            if name_b not in self.students:
                self.students[name_b] = Student(
                    name=name_b,
                    choice=epidosh_b,
                    gender=gender_b,
                    greek_knowledge=greek_b,
                    friends=[name_a],
                    locked=is_locked
                )
    
    def _load_from_single(self, sheet: Worksheet) -> None:
        """Διάβασμα μονών μαθητών από SINGLE sheet."""
        headers = self._parse_headers(sheet)
        
        required = ['ΟΝΟΜΑ', 'ΦΥΛΟ', 'ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ', 'ΕΠΙΔΟΣΗ']
        missing = [h for h in required if h not in headers]
        if missing:
            return
        
        for row_idx in range(2, sheet.max_row + 1):
            name = self._get_cell_value(sheet, row_idx, headers.get('ΟΝΟΜΑ'))
            if not name:
                continue
            
            if name in self.students:
                continue
            
            gender_col = headers.get('ΦΥΛΟ')
            greek_col = (headers.get('ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ') or 
                        headers.get('ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ') or
                        headers.get('ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ'))
            epidosh_col = headers.get('ΕΠΙΔΟΣΗ')
            locked_col = headers.get('LOCKED')
            
            gender = self._get_cell_value(sheet, row_idx, gender_col, 'Α')
            
            # Greek knowledge
            raw_greek = sheet.cell(row_idx, greek_col).value if greek_col else 'Ν'
            if raw_greek:
                greek_str = str(raw_greek).strip().upper()
                if greek_str.startswith('Ν') or greek_str.startswith('N'):
                    greek = 'Ν'
                elif greek_str.startswith('Ο') or greek_str.startswith('O'):
                    greek = 'Ο'
                else:
                    greek = 'Ν'
            else:
                greek = 'Ν'
            
            raw_epidosh = sheet.cell(row_idx, epidosh_col).value if epidosh_col else 1
            try:
                epidosh = int(raw_epidosh) if raw_epidosh else 1
            except:
                epidosh = 1
            
            locked_val = self._get_cell_value(sheet, row_idx, locked_col)
            is_locked = (locked_val == 'LOCKED')
            
            self.students[name] = Student(
                name=name,
                choice=epidosh,
                gender=gender,
                greek_knowledge=greek,
                friends=[],
                locked=is_locked
            )
    
    def calculate_spreads(self) -> Dict[str, int]:
        """Υπολογισμός spreads."""
        stats = self._get_team_stats()
        if not stats:
            return {'ep3': 0, 'boys': 0, 'girls': 0, 'greek_yes': 0}
        
        ep3_vals = [s['ep3'] for s in stats.values()]
        boys_vals = [s['boys'] for s in stats.values()]
        girls_vals = [s['girls'] for s in stats.values()]
        greek_yes_vals = [s['greek_yes'] for s in stats.values()]
        
        return {
            'ep3': max(ep3_vals) - min(ep3_vals),
            'boys': max(boys_vals) - min(boys_vals),
            'girls': max(girls_vals) - min(girls_vals),
            'greek_yes': max(greek_yes_vals) - min(greek_yes_vals)
        }
    
    def _get_team_stats(self) -> Dict:
        """Μέτρηση stats ανά τμήμα."""
        stats = {}
        for team_name, student_names in self.teams.items():
            boys = girls = greek_yes = greek_no = ep1 = ep2 = ep3 = 0
            
            for name in student_names:
                if name not in self.students:
                    continue
                s = self.students[name]
                
                if s.gender == 'Α':
                    boys += 1
                elif s.gender == 'Κ':
                    girls += 1
                
                if s.greek_knowledge in ['Ν', 'N']:
                    greek_yes += 1
                elif s.greek_knowledge in ['Ο', 'O']:
                    greek_no += 1
                
                if s.choice == 1:
                    ep1 += 1
                elif s.choice == 2:
                    ep2 += 1
                elif s.choice == 3:
                    ep3 += 1
            
            stats[team_name] = {
                'boys': boys, 'girls': girls,
                'greek_yes': greek_yes, 'greek_no': greek_no,
                'ep1': ep1, 'ep2': ep2, 'ep3': ep3
            }
        
        return stats
    
    def optimize(self, max_iterations: int = 100) -> Tuple[List[Dict], Dict]:
        """Asymmetric optimization."""
        applied_swaps = []
        
        print(f"🔄 Ξεκινά optimization (max {max_iterations} iterations)...")
        
        for iteration in range(max_iterations):
            spreads = self.calculate_spreads()
            
            if (spreads['ep3'] <= self.target_ep3 and
                spreads['boys'] <= self.target_gender and
                spreads['girls'] <= self.target_gender and
                spreads['greek_yes'] <= self.target_greek):
                print(f"✅ Targets επιτεύχθηκαν στο iteration {iteration}")
                break
            
            stats = self._get_team_stats()
            ep3_counts = {team: stats[team]['ep3'] for team in stats.keys()}
            
            max_team = max(ep3_counts.items(), key=lambda x: x[1])[0]
            min_team = min(ep3_counts.items(), key=lambda x: x[1])[0]
            
            if ep3_counts[max_team] - ep3_counts[min_team] <= self.target_ep3:
                print(f"✅ EP3 spread ≤ target στο iteration {iteration}")
                break
            
            all_swaps = self._generate_asymmetric_swaps(max_team, min_team)
            
            if not all_swaps:
                print(f"⚠️ Δεν βρέθηκαν άλλα swaps στο iteration {iteration}")
                break
            
            best_swap = self._select_best_swap(all_swaps)
            
            if not best_swap:
                print(f"⚠️ Δεν βρέθηκε best swap στο iteration {iteration}")
                break
            
            self._apply_swap(best_swap)
            applied_swaps.append(best_swap)
            
            if (iteration + 1) % 10 == 0:
                print(f"  Iteration {iteration + 1}: {len(applied_swaps)} swaps, spread_ep3={spreads['ep3']}")
        
        final_spreads = self.calculate_spreads()
        print(f"✅ Optimization ολοκληρώθηκε: {len(applied_swaps)} swaps")
        print(f"   Final spreads: EP3={final_spreads['ep3']}, Boys={final_spreads['boys']}, Girls={final_spreads['girls']}, Greek={final_spreads['greek_yes']}")
        
        return applied_swaps, final_spreads
    
    def _generate_asymmetric_swaps(self, max_team: str, min_team: str) -> List[Dict]:
        """Γέννηση asymmetric swaps."""
        swaps = []
        
        max_solos_ep3 = self._get_solos_with_ep3(max_team)
        max_pairs_ep3 = self._get_pairs_with_ep3(max_team)
        min_solos_non_ep3 = self._get_solos_without_ep3(min_team)
        min_pairs_non_ep3 = self._get_pairs_without_ep3(min_team)
        
        # P1: Solo(ep3)↔Solo(ep1/2) - same gender + greek
        for solo_max in max_solos_ep3:
            for solo_min in min_solos_non_ep3:
                if (solo_max['student'].gender == solo_min['student'].gender and
                    solo_max['student'].greek_knowledge == solo_min['student'].greek_knowledge):
                    
                    improvement = self._calc_asymmetric_improvement(
                        max_team, [solo_max['name']],
                        min_team, [solo_min['name']]
                    )
                    
                    if improvement['improves']:
                        swaps.append({
                            'type': 'Solo(ep3)↔Solo(ep1/2)-P1',
                            'from_team': max_team,
                            'students_out': [solo_max['name']],
                            'to_team': min_team,
                            'students_in': [solo_min['name']],
                            'improvement': improvement,
                            'priority': 1
                        })
        
        # P2: Pair swaps
        for pair_max in max_pairs_ep3:
            for pair_min in min_pairs_non_ep3:
                if (pair_max['student_a'].gender == pair_min['student_a'].gender and
                    pair_max['student_b'].gender == pair_min['student_b'].gender and
                    pair_max['student_a'].greek_knowledge == pair_min['student_a'].greek_knowledge and
                    pair_max['student_b'].greek_knowledge == pair_min['student_b'].greek_knowledge):
                    
                    improvement = self._calc_asymmetric_improvement(
                        max_team, [pair_max['name_a'], pair_max['name_b']],
                        min_team, [pair_min['name_a'], pair_min['name_b']]
                    )
                    
                    if improvement['improves']:
                        swaps.append({
                            'type': 'Pair(ep3+X)↔Pair(ep1/2)-P2',
                            'from_team': max_team,
                            'students_out': [pair_max['name_a'], pair_max['name_b']],
                            'to_team': min_team,
                            'students_in': [pair_min['name_a'], pair_min['name_b']],
                            'improvement': improvement,
                            'priority': 2
                        })
        
        # P3: Relaxed solo swaps
        for solo_max in max_solos_ep3:
            for solo_min in min_solos_non_ep3:
                if solo_max['student'].gender == solo_min['student'].gender:
                    if solo_max['student'].greek_knowledge == solo_min['student'].greek_knowledge:
                        continue
                    
                    improvement = self._calc_asymmetric_improvement(
                        max_team, [solo_max['name']],
                        min_team, [solo_min['name']]
                    )
                    if improvement['improves']:
                        swaps.append({
                            'type': 'Solo(ep3)↔Solo(ep1/2)-P3',
                            'from_team': max_team,
                            'students_out': [solo_max['name']],
                            'to_team': min_team,
                            'students_in': [solo_min['name']],
                            'improvement': improvement,
                            'priority': 3
                        })
        
        return swaps
    
    def _get_solos_with_ep3(self, team_name: str) -> List[Dict]:
        solos = []
        student_names = self.teams[team_name]
        for name in student_names:
            if name not in self.students:
                continue
            student = self.students[name]
            if student.locked or student.choice != 3:
                continue
            has_friend = any(f in student_names for f in student.friends)
            if not has_friend:
                solos.append({'name': name, 'student': student})
        return solos
    
    def _get_pairs_with_ep3(self, team_name: str) -> List[Dict]:
        pairs = []
        processed = set()
        student_names = self.teams[team_name]
        for name_a in student_names:
            if name_a in processed or name_a not in self.students:
                continue
            student_a = self.students[name_a]
            if student_a.locked:
                continue
            for name_b in student_names:
                if name_b == name_a or name_b in processed or name_b not in self.students:
                    continue
                student_b = self.students[name_b]
                if student_b.locked:
                    continue
                if name_b in student_a.friends or name_a in student_b.friends:
                    if student_a.choice == 3 or student_b.choice == 3:
                        pairs.append({
                            'name_a': name_a, 'name_b': name_b,
                            'student_a': student_a, 'student_b': student_b,
                            'ep_combo': f"{student_a.choice},{student_b.choice}"
                        })
                        processed.add(name_a)
                        processed.add(name_b)
                        break
        return pairs
    
    def _get_solos_without_ep3(self, team_name: str) -> List[Dict]:
        solos = []
        student_names = self.teams[team_name]
        for name in student_names:
            if name not in self.students:
                continue
            student = self.students[name]
            if student.locked or student.choice == 3:
                continue
            has_friend = any(f in student_names for f in student.friends)
            if not has_friend:
                solos.append({'name': name, 'student': student})
        return solos
    
    def _get_pairs_without_ep3(self, team_name: str) -> List[Dict]:
        pairs = []
        processed = set()
        student_names = self.teams[team_name]
        for name_a in student_names:
            if name_a in processed or name_a not in self.students:
                continue
            student_a = self.students[name_a]
            if student_a.locked:
                continue
            for name_b in student_names:
                if name_b == name_a or name_b in processed or name_b not in self.students:
                    continue
                student_b = self.students[name_b]
                if student_b.locked:
                    continue
                if name_b in student_a.friends or name_a in student_b.friends:
                    if student_a.choice != 3 and student_b.choice != 3:
                        pairs.append({
                            'name_a': name_a, 'name_b': name_b,
                            'student_a': student_a, 'student_b': student_b,
                            'ep_combo': f"{student_a.choice},{student_b.choice}"
                        })
                        processed.add(name_a)
                        processed.add(name_b)
                        break
        return pairs
    
    def _calc_asymmetric_improvement(self, team_high: str, names_out: List[str],
                                      team_low: str, names_in: List[str]) -> Dict:
        """Υπολογισμός improvement."""
        stats_before = self._get_team_stats()
        stats_after = {k: v.copy() for k, v in stats_before.items()}
        
        for name in names_out:
            if name in self.students:
                s = self.students[name]
                if s.choice == 3: stats_after[team_high]['ep3'] -= 1
                if s.gender == 'Α': stats_after[team_high]['boys'] -= 1
                elif s.gender == 'Κ': stats_after[team_high]['girls'] -= 1
                if s.greek_knowledge in ['Ν', 'N']: stats_after[team_high]['greek_yes'] -= 1
        
        for name in names_in:
            if name in self.students:
                s = self.students[name]
                if s.choice == 3: stats_after[team_high]['ep3'] += 1
                if s.gender == 'Α': stats_after[team_high]['boys'] += 1
                elif s.gender == 'Κ': stats_after[team_high]['girls'] += 1
                if s.greek_knowledge in ['Ν', 'N']: stats_after[team_high]['greek_yes'] += 1
        
        for name in names_in:
            if name in self.students:
                s = self.students[name]
                if s.choice == 3: stats_after[team_low]['ep3'] -= 1
                if s.gender == 'Α': stats_after[team_low]['boys'] -= 1
                elif s.gender == 'Κ': stats_after[team_low]['girls'] -= 1
                if s.greek_knowledge in ['Ν', 'N']: stats_after[team_low]['greek_yes'] -= 1
        
        for name in names_out:
            if name in self.students:
                s = self.students[name]
                if s.choice == 3: stats_after[team_low]['ep3'] += 1
                if s.gender == 'Α': stats_after[team_low]['boys'] += 1
                elif s.gender == 'Κ': stats_after[team_low]['girls'] += 1
                if s.greek_knowledge in ['Ν', 'N']: stats_after[team_low]['greek_yes'] += 1
        
        ep3_before = max(s['ep3'] for s in stats_before.values()) - min(s['ep3'] for s in stats_before.values())
        ep3_after = max(s['ep3'] for s in stats_after.values()) - min(s['ep3'] for s in stats_after.values())
        
        boys_before = max(s['boys'] for s in stats_before.values()) - min(s['boys'] for s in stats_before.values())
        boys_after = max(s['boys'] for s in stats_after.values()) - min(s['boys'] for s in stats_after.values())
        
        girls_before = max(s['girls'] for s in stats_before.values()) - min(s['girls'] for s in stats_before.values())
        girls_after = max(s['girls'] for s in stats_after.values()) - min(s['girls'] for s in stats_after.values())
        
        greek_before = max(s['greek_yes'] for s in stats_before.values()) - min(s['greek_yes'] for s in stats_before.values())
        greek_after = max(s['greek_yes'] for s in stats_after.values()) - min(s['greek_yes'] for s in stats_after.values())
        
        delta_ep3 = ep3_before - ep3_after
        delta_boys = boys_before - boys_after
        delta_girls = girls_before - girls_after
        delta_greek = greek_before - greek_after
        
        improves = delta_ep3 > 0 or (delta_ep3 == 0 and (delta_boys > 0 or delta_girls > 0 or delta_greek > 0))
        
        return {
            'improves': improves,
            'delta_ep3': delta_ep3,
            'delta_boys': delta_boys,
            'delta_girls': delta_girls,
            'delta_greek': delta_greek,
            'ep3_before': ep3_before,
            'ep3_after': ep3_after
        }
    
    def _select_best_swap(self, swaps: List[Dict]) -> Optional[Dict]:
        if not swaps:
            return None
        
        swaps.sort(
            key=lambda x: (
                -x['improvement']['delta_ep3'],
                -(x['improvement']['delta_boys'] + x['improvement']['delta_girls']),
                -x['improvement']['delta_greek'],
                x['priority']
            )
        )
        
        return swaps[0]
    
    def _apply_swap(self, swap: Dict) -> None:
        from_team = swap['from_team']
        to_team = swap['to_team']
        students_out = swap['students_out']
        students_in = swap['students_in']
        
        for name in students_out:
            if name in self.teams[from_team]:
                self.teams[from_team].remove(name)
        
        for name in students_in:
            if name in self.teams[to_team]:
                self.teams[to_team].remove(name)
        
        for name in students_out:
            self.teams[to_team].append(name)
        
        for name in students_in:
            self.teams[from_team].append(name)
    
    def export_optimized_excel(self, applied_swaps: List[Dict], final_spreads: Dict, output_path: str) -> str:
        """Εξαγωγή optimized Excel."""
        wb = Workbook()
        wb.remove(wb.active)
        
        for team_name in sorted(self.teams.keys()):
            self._create_team_sheet(wb, team_name)
        
        self._create_statistics_sheet(wb, final_spreads)
        self._create_swaps_log_sheet(wb, applied_swaps)
        
        wb.save(output_path)
        wb.close()
        
        print(f"✅ Optimized Excel αποθηκεύτηκε: {output_path}")
        return output_path
    
    def _create_team_sheet(self, wb: Workbook, team_name: str) -> None:
        sheet = wb.create_sheet(team_name)
        
        headers = ['ΟΝΟΜΑ', 'ΦΥΛΟ', 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ', 'ΕΠΙΔΟΣΗ', 'ΦΙΛΟΙ']
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(1, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDEBF7', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row_idx = 2
        for name in sorted(self.teams[team_name]):
            if name not in self.students:
                continue
            
            student = self.students[name]
            
            # Normalize greek_knowledge
            greek_val = student.greek_knowledge
            if greek_val in ['N', 'n']:
                greek_val = 'Ν'
            elif greek_val in ['O', 'o']:
                greek_val = 'Ο'
            
            sheet.cell(row_idx, 1).value = student.name
            sheet.cell(row_idx, 2).value = student.gender
            sheet.cell(row_idx, 3).value = greek_val
            sheet.cell(row_idx, 4).value = student.choice
            sheet.cell(row_idx, 5).value = ', '.join(student.friends)
            
            for col in range(1, 6):
                sheet.cell(row_idx, col).alignment = Alignment(
                    horizontal='left' if col in [1,5] else 'center', 
                    vertical='center'
                )
            
            row_idx += 1
        
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 25
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 40
    
    def _create_statistics_sheet(self, wb: Workbook, spreads: Dict) -> None:
        sheet = wb.create_sheet('ΒΕΛΤΙΩΜΕΝΗ_ΣΤΑΤΙΣΤΙΚΗ')
        
        headers = ['Τμήμα', 'Σύνολο', 'Αγόρια', 'Κορίτσια', 
                   'Γνώση (ΝΑΙ)', 'Γνώση (ΟΧΙ)', 'Επ1', 'Επ2', 'Επ3']
        
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(1, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='C6E0B4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        stats = self._get_team_stats()
        row_idx = 2
        for team_name in sorted(self.teams.keys()):
            if team_name not in stats:
                continue
            s = stats[team_name]
            
            sheet.cell(row_idx, 1).value = team_name
            sheet.cell(row_idx, 2).value = len(self.teams[team_name])
            sheet.cell(row_idx, 3).value = s['boys']
            sheet.cell(row_idx, 4).value = s['girls']
            sheet.cell(row_idx, 5).value = s['greek_yes']
            sheet.cell(row_idx, 6).value = s['greek_no']
            sheet.cell(row_idx, 7).value = s['ep1']
            sheet.cell(row_idx, 8).value = s['ep2']
            sheet.cell(row_idx, 9).value = s['ep3']
            
            for col in range(1, 10):
                sheet.cell(row_idx, col).alignment = Alignment(horizontal='center', vertical='center')
            
            row_idx += 1
        
        row_idx += 2
        sheet.cell(row_idx, 1).value = 'ΤΕΛΙΚΑ SPREADS'
        sheet.cell(row_idx, 1).font = Font(bold=True, size=12)
        row_idx += 1
        
        summary_headers = ['Μετρική', 'Spread', 'Στόχος', 'Status']
        for col_idx, header in enumerate(summary_headers, start=1):
            cell = sheet.cell(row_idx, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFF2CC', fill_type='solid')
        row_idx += 1
        
        summary_data = [
            ('Spread Επίδοσης 3', spreads['ep3'], '≤ 3', '✅' if spreads['ep3'] <= 3 else '❌'),
            ('Spread Αγοριών', spreads['boys'], '≤ 4', '✅' if spreads['boys'] <= 4 else '❌'),
            ('Spread Κοριτσιών', spreads['girls'], '≤ 4', '✅' if spreads['girls'] <= 4 else '❌'),
            ('Spread Γνώσης', spreads['greek_yes'], '≤ 4', '✅' if spreads['greek_yes'] <= 4 else '❌')
        ]
        
        for label, value, target, status in summary_data:
            sheet.cell(row_idx, 1).value = label
            sheet.cell(row_idx, 2).value = value
            sheet.cell(row_idx, 3).value = target
            sheet.cell(row_idx, 4).value = status
            
            if '✅' in status:
                sheet.cell(row_idx, 2).fill = PatternFill(start_color='C6EFCE', fill_type='solid')
            else:
                sheet.cell(row_idx, 2).fill = PatternFill(start_color='FFC7CE', fill_type='solid')
            
            row_idx += 1
        
        for col in ['A', 'B', 'C', 'D']:
            sheet.column_dimensions[col].width = 20
    
    def _create_swaps_log_sheet(self, wb: Workbook, swaps: List[Dict]) -> None:
        sheet = wb.create_sheet('ΕΦΑΡΜΟΣΜΕΝΑ_SWAPS')
        
        headers = ['#', 'Τύπος', 'Από Τμήμα', 'Μαθητές OUT', 
                   'Προς Τμήμα', 'Μαθητές IN', 'Δ_ep3', 'Δ_φύλου', 'Δ_γνώσης', 'Priority']
        
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(1, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for idx, swap in enumerate(swaps, start=1):
            imp = swap['improvement']
            
            sheet.cell(idx + 1, 1).value = idx
            sheet.cell(idx + 1, 2).value = swap['type']
            sheet.cell(idx + 1, 3).value = swap['from_team']
            sheet.cell(idx + 1, 4).value = ', '.join(swap['students_out'])
            sheet.cell(idx + 1, 5).value = swap['to_team']
            sheet.cell(idx + 1, 6).value = ', '.join(swap['students_in'])
            sheet.cell(idx + 1, 7).value = f"+{imp['delta_ep3']}" if imp['delta_ep3'] > 0 else str(imp['delta_ep3'])
            sheet.cell(idx + 1, 8).value = f"+{imp['delta_boys'] + imp['delta_girls']}" if imp['delta_boys'] + imp['delta_girls'] > 0 else str(imp['delta_boys'] + imp['delta_girls'])
            sheet.cell(idx + 1, 9).value = f"+{imp['delta_greek']}" if imp['delta_greek'] > 0 else str(imp['delta_greek'])
            sheet.cell(idx + 1, 10).value = swap['priority']
            
            for col in range(1, 11):
                sheet.cell(idx + 1, col).alignment = Alignment(horizontal='center', vertical='center')
        
        for col, width in [('A',8),('B',25),('C',15),('D',35),('E',15),('F',35),('G',10),('H',10),('I',10),('J',10)]:
            sheet.column_dimensions[col].width = width
    
    # ==================== HELPERS ====================
    
    def _parse_headers(self, sheet: Worksheet) -> Dict[str, int]:
        """Normalization headers."""
        headers = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                raw_header = str(cell.value).strip()
                headers[raw_header] = col_idx
                normalized = raw_header.upper().replace(' ', '').replace('_', '')
                headers[normalized] = col_idx
        return headers
    
    def _parse_headers_fill(self, sheet: Worksheet) -> Dict[str, int]:
        """Parse headers για fill phase."""
        headers = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                header = str(cell.value).strip()
                headers[header] = col_idx
        return headers
    
    def _get_cell_value(self, sheet: Worksheet, row: int, col: Optional[int], default: str = '') -> str:
        if col is None:
            return default
        val = sheet.cell(row, col).value
        return str(val).strip() if val is not None else default


# ========== CLI ==========

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Βήμα 8: Fill + Optimize (χωρίς Streamlit)")
    sub = p.add_subparsers(dest="mode", required=True, help="Λειτουργία")

    # Fill mode
    p_fill = sub.add_parser("fill", help="Fill template με δεδομένα")
    p_fill.add_argument("--source", required=True, help="Excel με μαθητές (Παράδειγμα1.xlsx)")
    p_fill.add_argument("--template", required=True, help="Template με τμήματα")
    p_fill.add_argument("--out", required=True, help="Output path")

    # Optimize mode
    p_opt = sub.add_parser("optimize", help="Optimize filled Excel")
    p_opt.add_argument("--input", required=True, help="Filled Excel")
    p_opt.add_argument("--out", required=True, help="Output path")
    p_opt.add_argument("--max-iter", type=int, default=100, help="Max iterations (default: 100)")

    # All mode
    p_all = sub.add_parser("all", help="Fill + Optimize σε μία")
    p_all.add_argument("--source", required=True)
    p_all.add_argument("--template", required=True)
    p_all.add_argument("--out", required=True)
    p_all.add_argument("--max-iter", type=int, default=100, help="Max iterations")

    return p


def main(argv: Optional[List[str]] = None) -> int:
    argv = argv if argv is not None else sys.argv[1:]
    parser = _build_parser()
    args = parser.parse_args(argv)

    try:
        processor = UnifiedProcessor()

        if args.mode == "fill":
            print(f"📄 Mode: FILL")
            processor.read_source_data(args.source)
            processor.fill_target_excel(args.template, args.out)
            
            if processor.warnings:
                print(f"\n⚠️  {len(processor.warnings)} warnings:")
                for w in processor.warnings[:10]:
                    print(f"  • {w}")
            
            return 0

        elif args.mode == "optimize":
            print(f"📄 Mode: OPTIMIZE")
            print("⚠️  Warning: Για optimize χρειάζεται το source file. Χρησιμοποίησε mode 'all'")
            return 1

        elif args.mode == "all":
            print(f"📄 Mode: ALL (Fill + Optimize)")
            
            # Phase 1: Fill
            print("\n📋 Phase 1/2: Filling...")
            processor.read_source_data(args.source)
            temp_filled = args.out.replace('.xlsx', '_TEMP_FILLED.xlsx')
            processor.fill_target_excel(args.template, temp_filled)
            
            # Phase 2: Optimize
            print("\n🎯 Phase 2/2: Optimizing...")
            processor.load_filled_data(temp_filled)
            
            spreads_before = processor.calculate_spreads()
            print(f"\n📊 ΠΡΙΝ:")
            print(f"  EP3 spread: {spreads_before['ep3']}")
            print(f"  Boys spread: {spreads_before['boys']}")
            print(f"  Girls spread: {spreads_before['girls']}")
            print(f"  Greek spread: {spreads_before['greek_yes']}")
            
            swaps, spreads_after = processor.optimize(max_iterations=args.max_iter)
            
            print(f"\n📊 ΜΕΤΑ:")
            print(f"  EP3 spread: {spreads_after['ep3']} {'✅' if spreads_after['ep3'] <= 3 else '❌'}")
            print(f"  Boys spread: {spreads_after['boys']} {'✅' if spreads_after['boys'] <= 4 else '❌'}")
            print(f"  Girls spread: {spreads_after['girls']} {'✅' if spreads_after['girls'] <= 4 else '❌'}")
            print(f"  Greek spread: {spreads_after['greek_yes']} {'✅' if spreads_after['greek_yes'] <= 4 else '❌'}")
            
            processor.export_optimized_excel(swaps, spreads_after, args.out)
            
            # Cleanup temp file
            Path(temp_filled).unlink(missing_ok=True)
            
            print(f"\n🎉 Ολοκληρώθηκε! Swaps: {len(swaps)}")
            
            if processor.warnings:
                print(f"\n⚠️  {len(processor.warnings)} warnings:")
                for w in processor.warnings[:10]:
                    print(f"  • {w}")
            
            return 0

    except FileNotFoundError as e:
        print(f"❌ Σφάλμα: Δεν βρέθηκε αρχείο - {e}", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"❌ Απρόσμενο σφάλμα: {e}", file=sys.stderr)
        return 2